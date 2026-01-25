from playwright.sync_api import sync_playwright
from openpyxl import Workbook
import os
import logging
import traceback
import sys
from datetime import datetime, timedelta

# 设置Django环境
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'myproject.settings')

import django
django.setup()

from mvp.models import ZhihuQuestion

# 配置日志
logger = logging.getLogger(__name__)

def searching(keyword, base_dir=None):
    """
    搜索知乎问题并保存到Excel文件
    
    Args:
        keyword (str): 搜索关键词
        base_dir (str, optional): 项目基础目录,用于定位文件
    """
    try:
        # 如果没有提供base_dir,则使用当前工作目录
        if base_dir is None:
            base_dir = os.getcwd()
            
        # 构建文件路径
        stealth_js_path = os.path.join(base_dir, 'stealth.min.js')
        excel_file_path = os.path.join(base_dir, 'q.xlsx')
        
        logger.info(f"开始搜索关键词: {keyword}")
        logger.info(f"基础目录: {base_dir}")
        logger.info(f"Stealth.js路径: {stealth_js_path}")
        logger.info(f"Excel文件路径: {excel_file_path}")
        
        # 检查stealth.min.js是否存在
        if not os.path.exists(stealth_js_path):
            logger.error(f"找不到文件: {stealth_js_path}")
            raise FileNotFoundError(f"找不到文件: {stealth_js_path}")
        
        # 创建工作簿对象
        wb = Workbook()
        sh = wb.active
        sh.append(["index", "question"])

        # 创建一个playwright上下文管理器
        with sync_playwright() as p:
            try:
                # 创建浏览器对象
                browser = p.chromium.launch_persistent_context(
                    # 指定本地Edge浏览器安装目录的绝对路径
                    executable_path=r'C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe',
                    # 指定本地Edge浏览器用户数据目录的绝对路径
                    user_data_dir=r"C:\Users\meiho\AppData\Local\Microsoft\Edge\User Data\AA",
                    # 开启有界面模式
                    headless=False,
                    # 设置窗口最大化
                    no_viewport=True,
                    args=['--start-maximized']
                )
                logger.info("浏览器启动成功")
            except Exception as e:
                logger.error(f"启动浏览器失败: {str(e)}")
                # 尝试无头模式作为备选方案
                try:
                    logger.info("尝试以无头模式启动浏览器")
                    browser = p.chromium.launch(headless=True)
                    logger.info("无头浏览器启动成功")
                except Exception as e2:
                    logger.error(f"启动无头浏览器也失败: {str(e2)}")
                    raise e  # 抛出原始错误

            page_list_aa = browser.pages
            logger.info(f'当前有 {len(page_list_aa)} 个页面')

            # 获取第一个页面对象
            page1 = page_list_aa[0] if page_list_aa else browser.new_page()
            
            # 防止被反爬
            try:
                with open(stealth_js_path, 'r', encoding="utf-8") as f:
                    js_data = f.read()
                page1.add_init_script(js_data)
                logger.info("成功加载stealth.js")
            except Exception as e:
                logger.warning(f"加载stealth.js失败: {str(e)}")

            page1.goto('https://www.zhihu.com')
            logger.info("访问知乎首页成功")
            
            page1.fill("//input[@type='text']", keyword)
            page1.click("//button[@aria-label='搜索']")
            page1.wait_for_timeout(1000)
            logger.info(f"搜索关键词 '{keyword}' 完成")
            
            # 减少滚动次数以提高性能
            for i in range(500):
                page1.wait_for_timeout(10)
                page1.evaluate(f'document.documentElement.scrollTop={(i + 1) * 500}')

            result_list = page1.locator("//span[@class='Highlight']").all()
            logger.info(f"找到 {len(result_list)} 个搜索结果")
            
            # 使用集合来存储已处理的问题,用于去重
            seen_questions = set()
            index = 1
            
            for res in result_list:
                try:
                    question_text = res.inner_text()
                    # 检查问题是否已经处理过
                    if question_text not in seen_questions:
                        seen_questions.add(question_text)
                        sh.append([index, question_text])
                        logger.info(f"第{index}个问题: {question_text[:50]}...")
                        index += 1
                    else:
                        logger.debug(f"跳过重复问题: {question_text[:50]}...")
                except Exception as e:
                    logger.warning(f"处理搜索结果时出错: {str(e)}")
                    continue

            # 保存文件
            wb.save(excel_file_path)
            logger.info(f"结果已保存到 {excel_file_path}")
            
            browser.close()
            logger.info("搜索完成")
            
    except Exception as e:
        logger.error(f"搜索过程中发生错误: {str(e)}")
        logger.error(traceback.format_exc())
        raise e

# ========================================
# 数据库适配函数
# ========================================

def check_zhihu_cache(keyword):
    """检查知乎问题缓存(7天)"""
    threshold = datetime.now() - timedelta(days=7)
    cached_count = ZhihuQuestion.objects.filter(
        keyword=keyword,
        created_at__gte=threshold
    ).count()
    return cached_count > 0, cached_count

def load_zhihu_questions(keyword):
    """从数据库加载缓存的问题"""
    threshold = datetime.now() - timedelta(days=7)
    questions = list(ZhihuQuestion.objects.filter(
        keyword=keyword,
        created_at__gte=threshold
    ).order_by('question_id').values('question_id', 'question_text'))
    return questions

def save_zhihu_questions_to_db(keyword, questions):
    """保存知乎问题到数据库"""
    # 先删除旧缓存
    ZhihuQuestion.objects.filter(keyword=keyword).delete()
    
    # 批量插入新数据
    question_objs = [
        ZhihuQuestion(
            keyword=keyword,
            question_id=q['question_id'],
            question_text=q['question_text']
        )
        for q in questions
    ]
    ZhihuQuestion.objects.bulk_create(question_objs, batch_size=100)

def searching_with_db(keyword, use_cache=True):
    """搜索知乎问题(数据库版本)"""
    try:
        # 缓存检查
        if use_cache:
            has_cache, count = check_zhihu_cache(keyword)
            if has_cache:
                logger.info(f"使用缓存: 找到 {count} 个问题")
                return load_zhihu_questions(keyword)
        
        # 执行搜索(调用原有逻辑)
        # 使用项目根目录作为base_dir
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        searching(keyword, base_dir=base_dir)
        
        # 读取 q.xlsx 并转换为数据库格式
        excel_file_path = os.path.join(base_dir, 'q.xlsx')
        from openpyxl import load_workbook
        wb = load_workbook(excel_file_path)
        sh = wb.active
        
        questions = []
        for i, row in enumerate(sh.iter_rows(min_row=2, values_only=True), 1):
            if row[0] and row[1]:
                questions.append({
                    'question_id': i,
                    'question_text': row[1]
                })
        
        # 保存到数据库
        save_zhihu_questions_to_db(keyword, questions)
        
        # 删除临时文件
        if os.path.exists(excel_file_path):
            os.remove(excel_file_path)
        
        logger.info(f"成功保存 {len(questions)} 个问题到数据库")
        
        return questions
        
    except Exception as e:
        logger.error(f"搜索知乎问题时出错: {str(e)}")
        logger.error(traceback.format_exc())
        raise e

if __name__ == "__main__":
    searching("新能源汽车")
